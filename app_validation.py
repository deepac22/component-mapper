


import re
import io
import fitz
import cv2
import numpy as np
import pytesseract
from pytesseract import Output
import ezdxf
import openpyxl
from openpyxl.styles import Font
import streamlit as st
import pandas as pd
from PIL import Image


CODE_PATTERN = re.compile(r'\bCO([A-Z]{1,4}\d+[A-Z]?)\b')
LABEL_PATTERN = re.compile(r'^[A-Z]{1,4}\d{1,4}[A-Z]?$')


# --------------------------------------------------------------
# SHARED OCR HELPER - used by PDF fallback and image files
# --------------------------------------------------------------
def ocr_grayscale_image(gray_image, grid=4):
    found = set()
    h, w = gray_image.shape
    overlap = 40
    for r in range(grid):
        for c in range(grid):
            y1 = max(0, r * h // grid - overlap)
            y2 = min(h, (r + 1) * h // grid + overlap)
            x1 = max(0, c * w // grid - overlap)
            x2 = min(w, (c + 1) * w // grid + overlap)
            tile = gray_image[y1:y2, x1:x2]
            for psm in [11, 6]:
                data = pytesseract.image_to_data(tile, output_type=Output.DICT, config=f'--psm {psm}')
                for i in range(len(data['text'])):
                    t = data['text'][i].strip().upper()
                    conf = int(data['conf'][i]) if data['conf'][i] != '-1' else -1
                    if conf > 35 and LABEL_PATTERN.match(t):
                        found.add(t)
    return found


# --------------------------------------------------------------
# FORMAT 1: PDF - real text first, OCR fallback
# --------------------------------------------------------------
def extract_from_pdf(file_bytes):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    all_codes = set()
    total_text_length = 0
    for page in doc:
        text = page.get_text()
        total_text_length += len(text)
        all_codes.update(CODE_PATTERN.findall(text))

    if len(all_codes) > 0:
        return sorted(all_codes), "text", None

    # FALLBACK: not every PDF uses the "CO" + designator netlist convention.
    # Some files just write the plain designator directly (e.g. "R100" not
    # "COR100"). Try matching bare designators as words in the text instead.
    if total_text_length >= 50:
        plain_codes = set()
        for page in doc:
            words = re.findall(r'\b[A-Za-z0-9]+\b', page.get_text())
            for w in words:
                w_upper = w.strip().upper()
                if LABEL_PATTERN.match(w_upper):
                    plain_codes.add(w_upper)
        if len(plain_codes) > 0:
            return sorted(plain_codes), "text-plain", None

    if total_text_length < 50:
        ocr_found = set()
        for page in doc:
            pix = page.get_pixmap(dpi=400)
            img_array = np.frombuffer(pix.tobytes("png"), dtype=np.uint8)
            img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            ocr_found |= ocr_grayscale_image(gray)
        if len(ocr_found) == 0:
            return [], None, "No embedded text, and OCR found no component labels either."
        return sorted(ocr_found), "ocr", None

    return [], None, "PDF has text, but no 'CO' + designator pattern was found."


# --------------------------------------------------------------
# FORMAT 2: DXF - CAD text-exchange format, real text entities
# --------------------------------------------------------------
def extract_from_dxf(file_bytes):
    try:
        text_stream = io.StringIO(file_bytes.decode('utf-8', errors='ignore'))
        doc = ezdxf.read(text_stream)
    except Exception as e:
        return [], f"Couldn't read this DXF file. ({e})"

    msp = doc.modelspace()
    found = set()

    for entity in msp.query('TEXT MTEXT ATTRIB'):
        try:
            text = entity.dxf.text if entity.dxftype() != 'MTEXT' else entity.text
        except Exception:
            continue
        if text and LABEL_PATTERN.match(text.strip().upper()):
            found.add(text.strip().upper())

    if len(found) == 0:
        return [], (
            "No component labels found in this DXF's text entities. "
            "The labels may be on a layer this parser didn't check, "
            "or drawn as geometry instead of text."
        )
    return sorted(found), None


# --------------------------------------------------------------
# FORMAT 3: Images (PNG/JPG/TIFF/BMP) - OCR only, no text layer exists
# --------------------------------------------------------------
def extract_from_image(file_bytes):
    try:
        pil_image = Image.open(io.BytesIO(file_bytes)).convert("RGB")
        img_array = np.array(pil_image)
        gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
    except Exception as e:
        return [], f"Couldn't open this image file. ({e})"

    found = ocr_grayscale_image(gray)
    if len(found) == 0:
        return [], "OCR couldn't find any component labels in this image."
    return sorted(found), None


# --------------------------------------------------------------
# MAIN DISPATCH - routes to the right extractor based on file type
# --------------------------------------------------------------
def extract_component_codes(uploaded_file):
    filename = uploaded_file.name.lower()
    file_bytes = uploaded_file.read()

    if filename.endswith('.pdf'):
        codes, method, error = extract_from_pdf(file_bytes)
        return codes, method, error

    elif filename.endswith('.dxf'):
        codes, error = extract_from_dxf(file_bytes)
        return codes, "dxf-text", error

    elif filename.endswith(('.png', '.jpg', '.jpeg', '.tif', '.tiff', '.bmp')):
        codes, error = extract_from_image(file_bytes)
        return codes, "ocr", error

    elif filename.endswith('.dwg'):
        return [], None, (
            "**.dwg files are not supported.** This is AutoCAD's proprietary "
            "binary format - there's no reliable free library to read it. "
            "Please re-export/save the drawing as **PDF or DXF** from your "
            "CAD tool and upload that instead."
        )

    else:
        return [], None, f"Unsupported file type: {filename.split('.')[-1] if '.' in filename else 'unknown'}"


# --------------------------------------------------------------
# REFERENCE TABLE - dynamic category matching (from previous fix)
# --------------------------------------------------------------
def load_reference_table(xlsx_bytes):
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        sheet = workbook.active
    except Exception as e:
        return [], f"Couldn't open this Excel file. ({e})"

    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    if not rows:
        return [], "This reference file appears to be empty."

    header_row = [str(h).strip().lower() if h else "" for h in rows[0]]
    code_col, desc_col = None, None
    for i, h in enumerate(header_row):
        if "code" in h and code_col is None:
            code_col = i
        if "desc" in h:
            desc_col = i
    if code_col is None:
        code_col = 0
    if desc_col is None:
        desc_col = 2 if len(header_row) > 2 else len(header_row) - 1

    entries = []
    for row in rows[1:]:
        if len(row) <= max(code_col, desc_col):
            continue
        code, desc = row[code_col], row[desc_col]
        if code and desc:
            entries.append((code, str(desc)))

    if not entries:
        return [], "Couldn't find usable code/description rows. Expected headers containing 'Code' and 'Description'."

    return entries, None


def get_prefix(component_code):
    match = re.match(r'^([A-Z]+)\d', component_code)
    return match.group(1) if match else component_code


PREFIX_TO_KEYWORDS = {
    'R': ['RESISTOR'], 'C': ['CAPACITOR'], 'L': ['INDUCTOR'],
    'D': ['DIODE', 'RECTIFIER'], 'Q': ['TRANSISTOR'], 'J': ['CONNECTOR'],
    'SW': ['SWITCH'], 'F': ['FUSE'], 'NT': ['ANTENNA'], 'ANT': ['ANTENNA'],
    'X': ['CRYSTAL', 'OSCILLATOR'], 'Y': ['CRYSTAL', 'OSCILLATOR'],
}
NEEDS_REVIEW_PREFIXES = {'U', 'MP', 'TP', 'PTH', 'FD', 'FILT', 'DMC', 'Z', 'P', 'A'}


def find_match_in_reference(keywords, reference_entries):
    """Searches the uploaded reference file for the most GENERIC match -
    the description with the fewest extra/unrelated words - so 'CAPACITOR'
    beats 'Super Capacitor Module'."""
    best_match = None
    best_extra_word_count = None
    for keyword in keywords:
        for code, description in reference_entries:
            words = re.findall(r'[A-Za-z]+', description.upper())
            keyword_word_matches = [w for w in words if keyword in w]
            if not keyword_word_matches:
                continue
            extra_words = len(words) - len(keyword_word_matches)
            if best_extra_word_count is None or extra_words < best_extra_word_count:
                best_extra_word_count = extra_words
                best_match = (code, description)
        if best_match:
            break
    return best_match


def map_component(component_code, reference_entries):
    prefix = get_prefix(component_code)
    if prefix in PREFIX_TO_KEYWORDS:
        match = find_match_in_reference(PREFIX_TO_KEYWORDS[prefix], reference_entries)
        if match:
            code, description = match
            return code, description, "OK"
        return "NOT IN FILE", f"No '{PREFIX_TO_KEYWORDS[prefix][0]}' category found in this reference file", "REVIEW"
    elif prefix in NEEDS_REVIEW_PREFIXES:
        return "NEEDS REVIEW", f"'{prefix}' has multiple possible categories - needs manual classification", "REVIEW"
    else:
        return "UNKNOWN PREFIX", f"'{prefix}' is not a recognized component family", "REVIEW"


def build_excel_bytes(component_codes, reference_entries):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    headers = ["Partno", "Component code in the drawing", "Mapping Code", "Description", "Status"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)
    for row_num, code in enumerate(component_codes, start=2):
        commodity_code, description, status = map_component(code, reference_entries)
        sheet.cell(row=row_num, column=1, value=row_num - 1)
        sheet.cell(row=row_num, column=2, value=code)
        sheet.cell(row=row_num, column=3, value=commodity_code)
        sheet.cell(row=row_num, column=4, value=description)
        sheet.cell(row=row_num, column=5, value=status)
    widths = {"A": 8, "B": 30, "C": 16, "D": 45, "E": 12}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------
# PAGE + STYLING
# --------------------------------------------------------------
st.set_page_config(page_title="CompMap | Flex", page_icon="◆", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&family=JetBrains+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'Space Grotesk', sans-serif; }
    .hero {
        padding: 2.2rem 2rem; border-radius: 16px;
        background: linear-gradient(135deg, #0091D5 0%, #005A8C 100%);
        margin-bottom: 1.8rem; box-shadow: 0 8px 32px rgba(0, 145, 213, 0.25);
        display: flex; justify-content: space-between; align-items: center;
    }
    .hero-left h1 { color: white; font-size: 2rem; font-weight: 700; margin: 0; letter-spacing: -0.02em; }
    .hero-left p { color: rgba(255,255,255,0.9); margin-top: 0.4rem; font-size: 0.95rem; }
    .badge {
        display: inline-block; background: rgba(255,255,255,0.2); color: white;
        padding: 3px 12px; border-radius: 20px; font-size: 0.7rem;
        font-family: 'JetBrains Mono', monospace; letter-spacing: 0.05em; margin-top: 0.6rem;
    }
    .flex-wordmark {
        font-family: 'Space Grotesk', sans-serif; font-weight: 700; font-size: 1.4rem;
        color: #0091D5; background: white; letter-spacing: 0.02em;
        padding: 8px 20px; border-radius: 8px;
    }
    div[data-testid="stMetric"] { background: #F0F8FC; border: 1px solid #B3E0F5; border-radius: 12px; padding: 1rem 1.2rem; }
    div[data-testid="stMetricValue"] { font-family: 'JetBrains Mono', monospace; color: #0091D5; }
    .stDataFrame { border-radius: 12px; overflow: hidden; }
    div[data-testid="stFileUploader"] { border: 1px dashed #0091D5; border-radius: 12px; padding: 0.5rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
    <div class="hero-left">
        <h1>◆ CompMap</h1>
        <p>Upload a PCBA assembly drawing (PDF, DXF, or image) and a reference file — get a mapping report.</p>
        <span class="badge">v2.1 · multi-format input · dynamic reference matching</span>
    </div>
    <div class="flex-wordmark">flex</div>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown("**📄 Assembly Drawing** — PDF, DXF, PNG, JPG, TIFF, or BMP")
    drawing_file = st.file_uploader(
        "Drawing", type=["pdf", "dxf", "png", "jpg", "jpeg", "tif", "tiff", "bmp", "dwg"],
        label_visibility="collapsed"
    )
with col2:
    st.markdown("**📋 Component Reference File**")
    ref_file = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed")

if drawing_file and ref_file:
    with st.spinner("Extracting and mapping components..."):
        codes, method, drawing_error = extract_component_codes(drawing_file)
        reference_entries, ref_error = load_reference_table(ref_file.read())

    if drawing_error:
        st.error(f"**Drawing issue:** {drawing_error}")
    if ref_error:
        st.error(f"**Reference file issue:** {ref_error}")

    if not drawing_error and not ref_error:
        if method == "ocr":
            st.warning(
                "⚠️ **This file has no embedded text** — results come from image-based "
                "reading (OCR), which is significantly less complete than text extraction. "
                "Treat this as a rough starting point, not a verified list."
            )
        elif method == "dxf-text":
            st.info("✓ Read directly from DXF text entities — same reliability as PDF text extraction.")
        elif method == "text-plain":
            st.warning(
                "⚠️ **This PDF doesn't use the standard 'CO' + designator netlist format** "
                "— results come from matching plain designator-shaped text instead (e.g. "
                "'R100', 'C205'). This is looser than the strict pattern and may include "
                "some false positives (random text that happens to look like a component "
                "code) or miss codes with unusual formatting. Spot-check against the "
                "drawing before treating this as final."
            )

        rows = []
        for i, code in enumerate(codes, start=1):
            mcode, desc, status = map_component(code, reference_entries)
            rows.append({
                "Partno": i, "Component code in the drawing": code,
                "Mapping Code": mcode, "Description": desc, "Status": status,
            })
        df = pd.DataFrame(rows)

        ok_count = int((df["Status"] == "OK").sum())
        review_count = int((df["Status"] == "REVIEW").sum())
        match_rate = round(100 * ok_count / len(df), 1) if len(df) else 0

        st.markdown("### Results")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total components", len(df))
        m2.metric("Confidently mapped", ok_count)
        m3.metric("Needs review", review_count)
        m4.metric("Match rate", f"{match_rate}%")

        tab1, tab2 = st.tabs(["All components", "Needs review"])
        with tab1:
            st.dataframe(df, use_container_width=True, height=400)
        with tab2:
            st.dataframe(df[df["Status"] == "REVIEW"], use_container_width=True, height=300)

        excel_bytes = build_excel_bytes(codes, reference_entries)
        st.download_button(
            "⬇ Download Excel Report", data=excel_bytes,
            file_name="Component_Mapping_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload both files above to generate your report.")
